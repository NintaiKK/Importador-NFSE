from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from tkinter import messagebox
import tkinter as tk
from tkinter import ttk
from tkinter import *
from tkinter.ttk import *
from openpyxl import Workbook
from openpyxl import load_workbook

janela = tk.Tk()
janela.resizable(0,0)
janela.geometry('800x500')
janela.title('Importador NFS-E')

a = tk.Frame()
b = tk.Frame()
c = tk.Frame()

def trabaio():
    
    servico = Service(ChromeDriverManager().install())
    navegador = webdriver.Chrome(service=servico)
    
    cnpj = cnpjEntry.get()
    senha = senhaEntry.get()
    diaI = entryDI.get()
    diaF = entryDf.get()

    navegador.get("https://nfse.canoas.rs.gov.br/autenticacao/login")
    navegador.find_element('xpath', '//*[@id="idUser"]').send_keys(cnpj)
    navegador.find_element('xpath', '//*[@id="idSenha"]').send_keys(senha)
    navegador.find_element('xpath', '//*[@id="btnLogin"]').click()

    try:
        element = WebDriverWait(navegador, 10).until(
            EC.visibility_of_element_located((By.XPATH, '//*[@id="MENU_NFSE"]'))
        )

        element.click()
    
    except Exception as e:
        print(f"Erro ao clicar no elemento: {e}")

    try:
        element1 = WebDriverWait(navegador, 10).until(
            EC.visibility_of_element_located((By.XPATH, '//*[@id="MENU_NFSE_EXPORTACAO/nfse/exportacao"]'))
        )

        element1.click()
    
    except Exception as e:
        print(f"Erro ao clicar no elemento: {e}")

    try:
        element3 = WebDriverWait(navegador, 10).until(
            EC.visibility_of_element_located((By.XPATH, '//*[@id="dataInicial"]'))
        )

        element3.click()

    except Exception as e:
        print(f"Erro ao clicar no elemento: {e}")

    navegador.find_element('xpath', '//*[@id="dataInicial"]').send_keys(diaI)

    navegador.find_element('xpath', '//*[@id="dataFinal"]').click()

    navegador.find_element('xpath', '//*[@id="dataFinal"]').send_keys(diaF)

    navegador.find_element('xpath', '//*[@id="btnExportarXml"]').click()

    try:
        element4 = WebDriverWait(navegador, 10).until(
            EC.visibility_of_element_located((By.XPATH, '//*[@id="tblMeusRelatorios"]/tbody/tr/td[4]/div/a[1]'))
        )

        element4.click()

    except Exception as e:
        print(f"Erro ao clicar no elemento: {e}")

    try:
        element11 = WebDriverWait(navegador, 10).until(
            EC.visibility_of_element_located((By.XPATH, '//*[@id="btnFecharModalRelatoriosAssincronos"]'))
        )

        element11.click()

    except Exception as e:
        print(f"Erro ao clicar no elemento: {e}")

    navegador.get("https://nfse.canoas.rs.gov.br/dmst/exportarNFSe")
    
    try:
        element7 = WebDriverWait(navegador, 10).until(
            EC.visibility_of_element_located((By.XPATH, '//*[@id="dataInicial"]'))
        )

        element7.click()

    except Exception as e:
        print(f"Erro ao clicar no elemento: {e}")

    navegador.find_element('xpath', '//*[@id="dataInicial"]').send_keys(diaI)

    navegador.find_element('xpath', '//*[@id="dataFinal"]').click()
    navegador.find_element('xpath', '//*[@id="dataFinal"]').send_keys(diaF)

    navegador.find_element('xpath', '//*[@id="btnAvancadoformPeriodo"]').click()
    
    try:
        element13 = WebDriverWait(navegador, 10).until(
            EC.visibility_of_element_located((By.XPATH, '//*[@id="tblNotas"]/tbody/tr/td[9]/div/div'))
        )

        valores = element13.text

    except Exception as e:
        print(f"Erro ao clicar no elemento: {e}")

    try:
        wb = load_workbook('Valores.xlsx')
        
    except FileNotFoundError:
        wb = Workbook()

    planilha = wb.active

    planilha.append([valores])

    wb.save('Valores.xlsx')

    try:
        element8 = WebDriverWait(navegador, 10).until(
            EC.visibility_of_element_located((By.XPATH, '//*[@id="formato"]'))
        )

        element8.click()

    except Exception as e:
        print(f"Erro ao clicar no elemento: {e}")

    try:
        element9 = WebDriverWait(navegador, 10).until(
            EC.visibility_of_element_located((By.XPATH, '//*[@id="formato"]/option[2]'))
        )

        element9.click()

    except Exception as e:
        print(f"Erro ao clicar no elemento: {e}")
    #aqui da pau

    navegador.quit()

lblVrd = tk.Label(
    master= a,
    height = 100,
    width = 45,
    text='Aprocont',
    bg='green')
lblVrd.pack()

lblSpcs = tk.Label(
    master = c,
    width = 10)
lblSpcs.pack()

lblCnpj = tk.Label(
    master = b,
    text = 'CNPJ')
lblCnpj.pack()

cnpjEntry = tk.Entry(
    master = b,
    width = 50)
cnpjEntry.pack()

lblSenha = tk.Label(
    master = b,
    text = 'Senha')
lblSenha.pack()

senhaEntry = tk.Entry(
    master = b,
    width = 50)
senhaEntry.pack()

lblDI = tk.Label(
    master = b,
    text = 'Data inicial')
lblDI.pack()

entryDI = tk.Entry(
    master = b,
    width = 50)
entryDI.pack()

lblDf = tk.Label(
    master = b,
    text = 'Data final')
lblDf.pack()

entryDf = tk.Entry(
    master = b,
    width = 50)
entryDf.pack()

lblFts = tk.Label(
    master = b,
    height = 1)
lblFts.pack()

btnLG = tk.Button(
    master = b,
    text = 'Importar',
    height = 1,
    command = trabaio)
btnLG.pack()

a.pack(side = LEFT)
c.pack(side = LEFT)
b.pack(side = LEFT)

janela.mainloop()

                                  
